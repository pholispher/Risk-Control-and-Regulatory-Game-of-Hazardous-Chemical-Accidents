# -*- coding: utf-8 -*-
"""
将《首批重点监管的危险化学品名录》和《危险化学品目录2015》两个.doc文件
转换为结构化Excel表格。

文档格式：
- 行内字段由 \\x07 (BEL, ASCII 7) 字符分隔
- 行由 \\r (CR, ASCII 13) 分隔
- 字段典型模式: \\r\\x07<value>\\r\\x07<value>...
"""

import os
import sys
import re
import win32com.client
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

sys.stdout.reconfigure(encoding='utf-8')

TASK_DIR = r"D:\Study\Chemicals\task"


def extract_text_from_doc(doc_path):
    """使用Word COM提取.doc文件中的纯文本"""
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    try:
        doc = word.Documents.Open(doc_path)
        text = doc.Content.Text
        doc.Close()
        return text
    finally:
        word.Quit()


def parse_bel_data(text, num_data_fields):
    """
    通用解析函数。

    文档使用 \\r\\x07 来分隔字段。
    通过 split('\\r\\x07') 得到所有 token，然后：
    - 跳过表头 token (如 "序号", "化学品名称", "品名" 等)
    - 当遇到纯数字 token 时，开始一条新记录
    - 收集后续 num_data_fields 个 token 作为该记录的字段
    """
    # 按 \r\x07 分割文本（而不是单独按 \r 或 \x07）
    tokens = text.split('\r\x07')
    # 清理每个 token（去除前后的 \r, \x07, 空白等）
    tokens = [t.strip('\r\x07 \t\x0c') for t in tokens]

    results = []
    i = 0
    while i < len(tokens):
        token = tokens[i]

        # 跳过空 token
        if not token:
            i += 1
            continue

        # 跳过表头关键词
        if token in ('序号', '化学品名称', '别名', 'CAS号', '品名', '备注',
                     '附件', '危险化学品目录', '说明', '首批重点监管的危险化学品名录'):
            i += 1
            continue

        # 检测序号（纯数字，1-4位）
        if not re.match(r'^\d{1,4}$', token):
            i += 1
            continue

        # 开始一条记录
        seq = int(token)
        record_tokens = [token]  # 序号
        i += 1

        # 收集后续 num_data_fields 个非空 token
        collected = 0
        while i < len(tokens) and collected < num_data_fields:
            t = tokens[i]
            # 如果遇到下一个序号，停止
            if re.match(r'^\d{1,4}$', t) and collected > 0:
                # 检查这是否像真的序号（不能是CAS号的中间部分）
                # 如果已收集到至少1个字段且下一个token是数字，可能是漏了字段
                # 这种情况下当作数据继续收集
                if collected >= num_data_fields - 1:
                    break
            record_tokens.append(t if t else '')
            collected += 1
            i += 1

        # 填充不足的字段
        while len(record_tokens) < num_data_fields + 1:
            record_tokens.append('')

        results.append(record_tokens)

    return results


def parse_list1(text):
    """
    解析《首批重点监管的危险化学品名录》
    字段: 序号, 化学品名称, 别名, CAS号 (4个数据字段)
    """
    raw = parse_bel_data(text, 4)
    entries = []
    for fields in raw:
        entry = {
            '序号': fields[0],
            '化学品名称': fields[1],
            '别名': fields[2],
            'CAS号': fields[3],
        }
        entries.append(entry)
    return entries


def parse_list2(text):
    """
    解析《危险化学品目录2015》
    字段: 序号, 品名, 别名, CAS号, 备注 (5个数据字段)
    """
    raw = parse_bel_data(text, 5)
    entries = []
    for fields in raw:
        entry = {
            '序号': fields[0],
            '品名': fields[1],
            '别名': fields[2],
            'CAS号': fields[3],
            '备注': fields[4],
        }
        entries.append(entry)
    return entries


def create_excel(entries, output_path, sheet_title, columns, col_widths=None):
    """创建格式化的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # 标题样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 数据样式
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, entry in enumerate(entries, 2):
        for col_idx, key in enumerate(columns, 1):
            value = entry.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 1:  # 序号居中
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment

    # 设置列宽
    if col_widths:
        from openpyxl.utils import get_column_letter
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return len(entries)


def main():
    # 文件路径
    doc1_path = os.path.join(TASK_DIR, "首批重点监管的危险化学品名录 .doc")
    doc2_path = os.path.join(TASK_DIR, "危险化学品目录2015.doc")

    # 输出路径
    xlsx1_path = os.path.join(TASK_DIR, "首批重点监管的危险化学品名录.xlsx")
    xlsx2_path = os.path.join(TASK_DIR, "危险化学品目录2015.xlsx")

    print("=" * 60)
    print("Processing documents...")
    print("=" * 60)

    # 处理第一个文档
    print("\n[1/4] Reading doc1...")
    text1 = extract_text_from_doc(doc1_path)
    print(f"    Extracted: {len(text1)} chars, {text1.count(chr(7))} BEL chars")

    print("[2/4] Parsing doc1...")
    entries1 = parse_list1(text1)
    print(f"    Parsed: {len(entries1)} records")

    # 处理第二个文档
    print("\n[3/4] Reading doc2...")
    text2 = extract_text_from_doc(doc2_path)
    print(f"    Extracted: {len(text2)} chars, {text2.count(chr(7))} BEL chars")

    print("[4/4] Parsing doc2...")
    entries2 = parse_list2(text2)
    print(f"    Parsed: {len(entries2)} records")

    # 生成Excel
    print("\n" + "=" * 60)
    print("Generating Excel files...")
    print("=" * 60)

    cols1 = ['序号', '化学品名称', '别名', 'CAS号']
    count1 = create_excel(entries1, xlsx1_path, '首批重点监管危险化学品', cols1,
                          col_widths=[8, 28, 30, 18])
    print(f"\n  [OK] {os.path.basename(xlsx1_path)}")
    print(f"       {count1} records")

    cols2 = ['序号', '品名', '别名', 'CAS号', '备注']
    count2 = create_excel(entries2, xlsx2_path, '危险化学品目录2015', cols2,
                          col_widths=[8, 40, 50, 18, 8])
    print(f"\n  [OK] {os.path.basename(xlsx2_path)}")
    print(f"       {count2} records")

    # 打印前几条作为预览
    print("\n" + "=" * 60)
    print("Preview - Doc1 (first 5):")
    print("=" * 60)
    for e in entries1[:5]:
        print(f"  {str(e['序号']):>4s} | {e['化学品名称']:<20s} | {e['别名']:<25s} | {e['CAS号']}")

    print("\n" + "=" * 60)
    print("Preview - Doc2 (first 5):")
    print("=" * 60)
    for e in entries2[:5]:
        print(f"  {str(e['序号']):>4s} | {e['品名'][:25]:<25s} | {e['别名'][:30]:<30s} | {e['CAS号']:<15s} | {e['备注']}")

    # 验证
    print("\n" + "=" * 60)
    print("Validation:")
    print("=" * 60)
    print(f"\n  Doc1: Expected 60, Got {len(entries1)}")
    print(f"  Doc2: Expected ~2828, Got {len(entries2)}")

    # Doc2 最后几条
    if entries2:
        print("\n  Doc2 - Last 3:")
        for e in entries2[-3:]:
            print(f"  {str(e['序号']):>4s} | {e['品名'][:30]:<30s} | {e['别名'][:20]:<20s} | {e['CAS号']}")

    print("\nDone!")


if __name__ == '__main__':
    main()
