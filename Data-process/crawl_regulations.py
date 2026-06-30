# -*- coding: utf-8 -*-
"""
危险化学品法规/标准爬虫 (最终版)
================================
数据源: law.chemicalsafety.org.cn
三个模块: 法律法规(4781) / 国家标准(11421) / 行业标准(5980)
总计约 22,182 条记录

使用: python crawl_regulations.py
     在浏览器中登录，之后全自动抓取
"""

import os, sys, time, re
sys.stdout.reconfigure(encoding='utf-8')

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.edge.options import Options as EdgeOptions
from html import unescape

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ==================== 配置 ====================
TASK_DIR = r"D:\Study\Chemicals\task"
BASE_URL = "https://law.chemicalsafety.org.cn/compliance/guild/customer/RegulationCustomer.jsp?moduleId=2&libraryId=3382332752404776&type=regulation"

MODULES = [
    ("法律法规", "regulation"),
    ("国家标准", "countryTech"),
    ("行业标准", "industryTech"),
]

PAGE_SIZE = 50           # 每页条数 (10/20/50, 50最快)
PAGE_DELAY = 2.0         # 翻页等待（秒）
DETAIL_DELAY = 1.5       # 详情等待
MAX_PAGES = None         # None=全部, 数字=限制页数(测试)

OUTPUT_LIST = os.path.join(TASK_DIR, "法规列表.xlsx")
OUTPUT_DETAIL = os.path.join(TASK_DIR, "法规详情.xlsx")

# ==================== 浏览器 ====================

def create_driver():
    options = EdgeOptions()
    options.add_argument('--no-sandbox')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    return webdriver.Edge(options=options)


def wait_for_login(driver, timeout=120):
    print(">>> 请在 Edge 浏览器中登录 <<<")
    for i in range(timeout):
        time.sleep(1)
        try:
            if driver.find_elements(By.XPATH, "//div[text()='法律法规']") and \
               '短信快捷登录' not in driver.page_source[:3000]:
                print(f"[OK] 登录成功! ({i+1}秒)\n")
                return True
        except:
            pass
        if i % 15 == 14:
            print(f"  等待中... ({i+1}秒)")
    return False


# ==================== 列表抓取 ====================

def switch_module(driver, module_name):
    """点击模块Tab并设置大pageSize"""
    tabs = driver.find_elements(By.XPATH, f"//div[text()='{module_name}']")
    if not tabs:
        print(f"  [ERROR] 找不到Tab: {module_name}")
        return False
    driver.execute_script("arguments[0].click();", tabs[0])
    time.sleep(3)

    # 设置 pageSize
    try:
        sel = Select(driver.find_element(By.CSS_SELECTOR, 'select.pagination-page-list'))
        sel.select_by_visible_text(str(PAGE_SIZE))
        time.sleep(2)
        print(f"  pageSize = {PAGE_SIZE}")
    except Exception as e:
        print(f"  [WARN] 设置pageSize失败: {e}")
    return True


def get_page_info(driver):
    """获取总页数和总记录数"""
    html = driver.page_source
    pages = re.search(r'共\s*(\d+)\s*页', html)
    total = re.search(r'总共\s*(\d+)', html)
    return {
        'total_pages': int(pages.group(1)) if pages else 1,
        'total_records': int(total.group(1)) if total else 0,
    }


def parse_href(href):
    """解析 openDynAttachment*(...) 链接，支持不同函数变体"""
    if not href:
        return {}
    # 匹配所有 openDynAttachment* 函数
    m = re.search(r"openDynAttachment\w+\(([^)]+)\)", href)
    if m:
        args_str = m.group(1)
        # 提取所有带引号的参数
        args = re.findall(r'"([^"]*)"', args_str)
        if len(args) >= 2:
            return {'recordId': args[0], 'attId': args[1],
                    'isEnglish': args[2] if len(args) > 2 else '',
                    'attId2': args[3] if len(args) > 3 else ''}
    return {}


def extract_page(driver):
    """提取当前页 datagrid 数据"""
    items = []
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, 'table.datagrid-btable tr.datagrid-row')
        for row in rows:
            tds = row.find_elements(By.TAG_NAME, 'td')
            if len(tds) < 3:
                continue

            cells = []
            detail = {}
            for td in tds:
                cells.append(td.text.strip())
                # 提取详情链接 (每个td都有相同的链接)
                links = td.find_elements(By.TAG_NAME, 'a')
                for a in links:
                    href = a.get_attribute('href') or ''
                    if 'openDynAttachment' in href and not detail:
                        detail = parse_href(href)

            items.append({
                'cells': cells,
                'recordId': detail.get('recordId', ''),
                'attId': detail.get('attId', ''),
                'isEnglish': detail.get('isEnglish', ''),
                'attId2': detail.get('attId2', ''),
            })
    except Exception as e:
        print(f"    [WARN] extract_page: {e}")
    return items


def go_to_page(driver, page_num):
    """翻到指定页: 输入页码 + Enter"""
    try:
        page_input = driver.find_element(By.CSS_SELECTOR, 'input.pagination-num')
        page_input.clear()
        page_input.send_keys(str(page_num))
        page_input.send_keys(Keys.ENTER)
        time.sleep(PAGE_DELAY)
        return True
    except Exception as e:
        print(f"    [WARN] go_to_page: {e}")
        return False


def scrape_module(driver, module_name):
    """抓取一个模块的全部列表"""
    print(f"\n{'='*60}")
    print(f"[模块] {module_name}")
    print(f"{'='*60}")

    if not switch_module(driver, module_name):
        return []

    info = get_page_info(driver)
    total_pages = info['total_pages']
    total_records = info['total_records']
    print(f"总记录: {total_records}, 总页数: {total_pages} (@{PAGE_SIZE}条/页)")

    pages_to_do = min(total_pages, MAX_PAGES) if MAX_PAGES else total_pages
    all_items = []

    for page in range(1, pages_to_do + 1):
        if page > 1:
            if not go_to_page(driver, page):
                print(f"[SKIP] 第{page}页翻页失败")
                continue

        items = extract_page(driver)

        # 去重相邻页
        if all_items and items and all_items[-1].get('recordId') and \
           all_items[-1]['recordId'] == items[0].get('recordId'):
            items = items[1:]

        all_items.extend(items)

        if page % 20 == 0 or page == pages_to_do:
            pct = page * 100 // pages_to_do
            print(f"  [{pct:3d}%] 第{page}/{pages_to_do}页 | {len(all_items)} 条")

    print(f"[完成] {module_name}: {len(all_items)} 条")
    return all_items


# ==================== 详情抓取 ====================

def scrape_detail(driver, item):
    """抓取单条详情"""
    rid = item.get('recordId', '')
    aid = item.get('attId', '')
    is_en = item.get('isEnglish', '')
    aid2 = item.get('attId2', '')
    title = item.get('cells', ['', '', ''])[2] if len(item.get('cells', [])) > 2 else ''

    if not rid:
        return {'title': title, 'recordId': '', 'content': '[无ID]'}

    try:
        # 检测用哪个函数: Regulation(4参数) vs Tech(2参数)
        if is_en or aid2:
            js = f"openDynAttachmentRegulation('{rid}','{aid}','{is_en}','{aid2}');"
        else:
            js = f"openDynAttachmentTech('{rid}','{aid}');"
        driver.execute_script(js)
        time.sleep(DETAIL_DELAY)

        content = ''

        # 弹出窗口 (artDialog)
        try:
            dlg = driver.find_element(By.CSS_SELECTOR, '.artDialog, div.ui-dialog-content')
            if dlg.is_displayed():
                iframes = dlg.find_elements(By.TAG_NAME, 'iframe')
                if iframes:
                    driver.switch_to.frame(iframes[0])
                    content = driver.find_element(By.TAG_NAME, 'body').text
                    driver.switch_to.default_content()
                else:
                    content = dlg.text
        except:
            pass

        # 新标签页
        if not content:
            handles = driver.window_handles
            if len(handles) > 1:
                driver.switch_to.window(handles[-1])
                time.sleep(1)
                try:
                    content = driver.find_element(By.TAG_NAME, 'body').text
                except:
                    pass
                driver.close()
                driver.switch_to.window(handles[0])

        # 关闭弹窗
        try:
            driver.execute_script(
                "if(typeof artDialog!=='undefined'){"
                "var d=artDialog.list;for(var i in d)d[i].close();}")
        except:
            pass
        try:
            close_btns = driver.find_elements(By.CSS_SELECTOR,
                '.artDialog-close, .ui-dialog-close')
            for btn in close_btns:
                if btn.is_displayed():
                    btn.click()
                    break
        except:
            pass

        if len(content) > 50000:
            content = content[:50000] + "\n\n[... 截断]"

        return {'title': title, 'recordId': rid, 'content': content or '[空]'}

    except Exception as e:
        return {'title': title, 'recordId': rid, 'content': f'[错误: {e}]'}


# ==================== Excel ====================

def save_list_excel(all_data, path):
    wb = Workbook(); ws = wb.active; ws.title = '法规列表'
    hf = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    df = Font(name='微软雅黑', size=10)
    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['模块', '序号', '文件/标准编号', '标题', '日期', '状态', '查看',
               '记录ID', '附件ID']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = bd

    row = 2
    for mod_name, items in all_data.items():
        for i, item in enumerate(items, 1):
            cells = item.get('cells', [])
            vals = [mod_name, i,
                    cells[1] if len(cells) > 1 else '',
                    cells[2] if len(cells) > 2 else '',
                    cells[3] if len(cells) > 3 else '',
                    cells[4] if len(cells) > 4 else '',
                    cells[5] if len(cells) > 5 else '',
                    item.get('recordId', ''),
                    item.get('attId', '')]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = df; c.border = bd
                c.alignment = ca if ci <= 2 else la
            row += 1

    for i, w in enumerate([12, 8, 35, 65, 15, 12, 8, 25, 25], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"\n[列表] {path} ({row-2} 行)")


def save_detail_excel(all_details, path):
    wb = Workbook(); ws = wb.active; ws.title = '法规详情'
    hf = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    df = Font(name='微软雅黑', size=10)
    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['模块', '序号', '标题', '记录ID', '正文']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = bd

    row = 2
    for mod_name, details in all_details.items():
        for i, d in enumerate(details, 1):
            for ci, v in enumerate([mod_name, i, d.get('title',''),
                                    d.get('recordId',''), d.get('content','')], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = df; c.border = bd
                c.alignment = ca if ci <= 2 else la
            row += 1

    for i, w in enumerate([12, 8, 50, 25, 90], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = 'A2'
    wb.save(path)
    print(f"[详情] {path} ({row-2} 行)")


# ==================== 主流程 ====================

def main():
    print("=" * 60)
    print("  危险化学品法规/标准爬虫")
    print(f"  pageSize={PAGE_SIZE}" +
          (f" MAX={MAX_PAGES}页(测试)" if MAX_PAGES else " 全量"))
    print("=" * 60)

    driver = create_driver()
    driver.maximize_window()
    driver.get(BASE_URL)

    if not wait_for_login(driver):
        print("[ERROR] 登录超时")
        driver.quit()
        return

    # === 阶段1: 列表 ===
    all_items = {}
    for mod_name, _ in MODULES:
        all_items[mod_name] = scrape_module(driver, mod_name)

    save_list_excel(all_items, OUTPUT_LIST)

    total = sum(len(v) for v in all_items.values())
    print(f"\n列表完成: {total} 条")
    for m, items in all_items.items():
        print(f"  {m}: {len(items)}")

    # === 阶段2: 详情 ===
    est_min = max(1, total * 1.5 / 60)
    print(f"\n抓取详情预计 ~{est_min:.0f} 分钟")

    # 无交互自动模式：总条数超过测试阈值才询问
    if MAX_PAGES and total < 100:
        choice = input("输入 y 抓取详情: ").strip().lower()
    else:
        print("自动跳过详情 (全量模式先确保列表正确)")
        choice = 'n'

    if choice == 'y':
        all_details = {}
        for mod_name, items in all_items.items():
            print(f"\n[详情] {mod_name} ({len(items)} 条)")
            details = []
            for i, item in enumerate(items, 1):
                detail = scrape_detail(driver, item)
                details.append(detail)
                if i % 100 == 0:
                    print(f"  {i}/{len(items)}")
            all_details[mod_name] = details
        save_detail_excel(all_details, OUTPUT_DETAIL)

    print("\n[完成] 5秒后关闭浏览器...")
    time.sleep(5)
    driver.quit()


if __name__ == '__main__':
    main()
